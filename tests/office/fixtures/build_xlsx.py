"""Generate Excel fixture files."""

from datetime import date
from pathlib import Path

from openpyxl import Workbook

HERE = Path(__file__).parent


def build_single_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    rows = [
        ["Apple", "Banana", "Cherry"],
        ["Dog", "Elephant", "Fox"],
        ["Hello", "World", "!"],
    ]
    for row in rows:
        ws.append(row)
    ws.append([1, 2, 3])
    wb.save(HERE / "single_sheet.xlsx")


def build_multi_sheet():
    wb = Workbook()
    wb.active.title = "Sheet A"
    wb.active["A1"] = "Hello"
    wb.active["B1"] = "World"

    s2 = wb.create_sheet("Sheet B")
    s2["A1"] = "Another text"
    s2["A2"] = "More text"

    s3 = wb.create_sheet("Sheet C")
    s3["A1"] = "Third"

    wb.save(HERE / "multi_sheet.xlsx")


def build_with_formulas():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Label"
    ws["A2"] = 10
    ws["A3"] = 20
    ws["A4"] = "=SUM(A2:A3)"
    ws["B1"] = "Also a label"
    wb.save(HERE / "with_formulas.xlsx")


def build_with_merged():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Merged Title"
    ws.merge_cells("A1:C1")
    ws["A2"] = "Col1"
    ws["B2"] = "Col2"
    ws["C2"] = "Col3"
    ws["A3"] = date(2026, 1, 1)
    ws["B3"] = 42
    ws["C3"] = "Some text"
    wb.save(HERE / "with_merged.xlsx")


if __name__ == "__main__":
    build_single_sheet()
    build_multi_sheet()
    build_with_formulas()
    build_with_merged()
    print("XLSX fixtures generated in:", HERE)
