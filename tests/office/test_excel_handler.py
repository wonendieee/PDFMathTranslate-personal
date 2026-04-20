"""Integration tests for XlsxFormatHandler."""
import asyncio
import shutil
from datetime import date
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook
from tests.office.helpers import MockTranslator
from translator.format.excel import XlsxFormatHandler

FIXTURES = Path(__file__).parent / "fixtures"


def _make_settings(output_dir):
    from translator.config.model import SettingsModel
    from translator.config.translate_engine_model import GoogleSettings

    s = SettingsModel(translate_engine_settings=GoogleSettings())
    s.translation.lang_in = "en"
    s.translation.lang_out = "zh"
    s.translation.qps = 4
    s.translation.pool_max_workers = 4
    s.translation.output = str(output_dir)
    return s


def test_single_sheet_strings_translated(tmp_path):
    src = FIXTURES / "single_sheet.xlsx"
    work = tmp_path / "single_sheet.xlsx"
    shutil.copy(src, work)

    mt = MockTranslator()
    settings = _make_settings(tmp_path)
    with patch("translator.format.excel.get_translator", return_value=mt):
        handler = XlsxFormatHandler()

        async def run():
            events = []
            async for ev in handler.translate(work, settings):
                events.append(ev)
            return events

        events = asyncio.run(run())

    assert events[-1]["type"] == "finish"
    out = Path(events[-1]["translate_result"].translated_path)
    assert out.exists()
    assert out.name == "single_sheet.zh.xlsx"

    wb = load_workbook(out)
    ws = wb.active
    assert ws["A1"].value == "[zh]Apple"
    assert ws["A4"].value == 1
    assert ws["B4"].value == 2


def test_formulas_preserved(tmp_path):
    src = FIXTURES / "with_formulas.xlsx"
    work = tmp_path / "with_formulas.xlsx"
    shutil.copy(src, work)

    mt = MockTranslator()
    settings = _make_settings(tmp_path)
    with patch("translator.format.excel.get_translator", return_value=mt):
        handler = XlsxFormatHandler()

        async def run():
            async for _ in handler.translate(work, settings):
                pass

        asyncio.run(run())

    out = tmp_path / "with_formulas.zh.xlsx"
    wb = load_workbook(out)
    ws = wb.active
    assert ws["A1"].value == "[zh]Label"
    assert ws["A4"].value == "=SUM(A2:A3)"
    assert ws["A2"].value == 10


def test_multi_sheet(tmp_path):
    src = FIXTURES / "multi_sheet.xlsx"
    work = tmp_path / "multi_sheet.xlsx"
    shutil.copy(src, work)

    mt = MockTranslator()
    settings = _make_settings(tmp_path)
    with patch("translator.format.excel.get_translator", return_value=mt):
        handler = XlsxFormatHandler()

        async def run():
            async for _ in handler.translate(work, settings):
                pass

        asyncio.run(run())

    out = tmp_path / "multi_sheet.zh.xlsx"
    wb = load_workbook(out)
    assert wb["Sheet A"]["A1"].value == "[zh]Hello"
    assert wb["Sheet B"]["A1"].value == "[zh]Another text"
    assert wb["Sheet C"]["A1"].value == "[zh]Third"


def test_merged_cells_and_date_preserved(tmp_path):
    src = FIXTURES / "with_merged.xlsx"
    work = tmp_path / "with_merged.xlsx"
    shutil.copy(src, work)

    mt = MockTranslator()
    settings = _make_settings(tmp_path)
    with patch("translator.format.excel.get_translator", return_value=mt):
        handler = XlsxFormatHandler()

        async def run():
            async for _ in handler.translate(work, settings):
                pass

        asyncio.run(run())

    out = tmp_path / "with_merged.zh.xlsx"
    wb = load_workbook(out)
    ws = wb.active
    assert ws["A1"].value == "[zh]Merged Title"
    assert "A1:C1" in [str(r) for r in ws.merged_cells.ranges]
    # openpyxl may deserialize as datetime; compare date component
    a3 = ws["A3"].value
    from datetime import datetime as _dt
    if isinstance(a3, _dt):
        assert a3.date() == date(2026, 1, 1)
    else:
        assert a3 == date(2026, 1, 1)
    assert ws["B3"].value == 42
    assert ws["C3"].value == "[zh]Some text"
