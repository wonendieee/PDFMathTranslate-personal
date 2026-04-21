"""Excel format handler (.xlsx) - full pipeline."""

from __future__ import annotations

import logging
import time
from collections.abc import AsyncGenerator
from dataclasses import dataclass
from pathlib import Path
from typing import TYPE_CHECKING

from translator.format.base import DocumentFormat
from translator.format.base import FormatHandler
from translator.office.batch_translator import OfficeBatchTranslator
from translator.office.text_utils import should_translate

if TYPE_CHECKING:
    from translator.config.model import SettingsModel

logger = logging.getLogger(__name__)


def get_translator(settings):
    """Lazy wrapper; tests can patch ``translator.format.excel.get_translator``."""
    from translator.engines import get_translator as _g

    return _g(settings)


@dataclass
class ExcelTranslateResult:
    """GUI 会对 mono_pdf_path 等字段调 .exists()，保持 Path | None 类型。"""

    original_path: str
    translated_path: str
    total_seconds: float
    mono_pdf_path: Path | None = None
    dual_pdf_path: Path | None = None
    auto_extracted_glossary_path: Path | None = None

    @property
    def original_pdf_path(self) -> str:
        return self.original_path


class XlsxFormatHandler(FormatHandler):
    """.xlsx handler. .xls (legacy binary) is NOT supported."""

    def get_format(self) -> DocumentFormat:
        return DocumentFormat.XLSX

    def detect_format(self, file_path: Path) -> bool:
        if not file_path.exists():
            return False
        if file_path.suffix.lower() != ".xlsx":
            return False
        try:
            import zipfile

            with zipfile.ZipFile(file_path, "r") as zf:
                return any(n.startswith("xl/") for n in zf.namelist())
        except Exception:
            return False

    def validate_file(self, file_path: Path) -> bool:
        if not self.detect_format(file_path):
            return False
        try:
            import openpyxl

            wb = openpyxl.load_workbook(file_path, read_only=True)
            _ = wb.sheetnames
            wb.close()
            return True
        except Exception as e:
            logger.debug(f".xlsx validation failed for {file_path}: {e}")
            return False

    async def translate(
        self,
        input_file: Path,
        settings: SettingsModel,
    ) -> AsyncGenerator[dict, None]:
        import openpyxl

        start = time.perf_counter()

        yield {"type": "progress", "overall_progress": 0.05, "stage": "reading"}
        wb = openpyxl.load_workbook(input_file)

        yield {"type": "progress", "overall_progress": 0.10, "stage": "collecting"}
        targets: list[tuple[str, str, str]] = []
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if not isinstance(v, str):
                        continue
                    if v.startswith("="):
                        continue
                    if not should_translate(v):
                        continue
                    targets.append((ws_name, cell.coordinate, v))

        translator_obj = get_translator(settings)
        qps = settings.translation.qps
        pool_workers = settings.translation.pool_max_workers or qps
        bt = OfficeBatchTranslator(
            translator_obj,
            qps=qps,
            max_workers=pool_workers,
        )
        texts = [t[2] for t in targets]
        if texts:
            translations = await bt.translate_batch(texts)
        else:
            translations = []
        yield {"type": "progress", "overall_progress": 0.85, "stage": "translating"}

        yield {"type": "progress", "overall_progress": 0.95, "stage": "writing"}
        for (ws_name, coord, _original), translated in zip(
            targets, translations, strict=True
        ):
            wb[ws_name][coord].value = translated

        output_dir = (
            Path(settings.translation.output)
            if settings.translation.output
            else input_file.parent
        )
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"{input_file.stem}.zh.xlsx"
        wb.save(output_path)

        elapsed = time.perf_counter() - start
        result = ExcelTranslateResult(
            original_path=str(input_file),
            translated_path=str(output_path),
            total_seconds=elapsed,
            mono_pdf_path=output_path,
        )

        yield {
            "type": "finish",
            "translate_result": result,
            "token_usage": {},
        }
